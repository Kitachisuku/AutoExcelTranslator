from openpyxl import load_workbook
from googletrans import Translator
import sys

# 数字を取り除く
def delNumber(number):
    num = number
    for i in range(0,9):
        num = num.replace(str(i),"")
    return num

# 文字を取り除く
def delAlphabet(alphabet):
    alpha = str(alphabet)
    alphabets = [chr(ord('a') + i).upper() for i in range(26)]
    for i in range(26):
        alpha = alpha.replace(alphabets[i],"")
    return alpha

args = sys.argv[1]

workbook = load_workbook(filename=args,read_only=False)

sheet = workbook[workbook.sheetnames[0]]

#   どこからどこまでが描いてる範囲なのか調べる
cells = sheet.dimensions.split(':')

print("Loading Cells...")

alphabets = [chr(ord('a') + i).upper() for i in range(26)]

for ia in range(0,26):
    for i in range(1,int(delAlphabet(cells[1]))):
        # A段落のi番目のセルの文字を取得
        celltext = sheet[str(alphabets[ia]) + str(i)].value
        # 何も書いてないのなら空白にする
        if celltext == None:
            celltext = ""
        else:
            #   翻訳
            translator = Translator(service_urls=['translate.googleapis.com'])
            while True:
                try:
                    transtext = translator.translate(celltext,dest="ja").text
                    sheet[alphabets[ia] + str(i)].value = transtext
                    print("Translate celltext to ",transtext)
                    break
                except Exception as e:
                    translator = Translator(service_urls=['translate.googleapis.com'])

print("done.")
workbook.save(args + ".translated.xlsx")
workbook.close()